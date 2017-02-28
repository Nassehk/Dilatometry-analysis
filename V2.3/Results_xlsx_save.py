# -*- coding: utf-8 -*-
"""
Created on Wed Dec 30 09:07:26 2015

Cross platform (Windows and Linux compatible)
@author: nasseh
"""
global sr, reg_order, interval
#this sript saves the results of calculation in the xlsx format
def saver(file_list,time_cond_master, temp_cond_master, dil_cond_master, \
result_master_time, result_master_temp, result_master_dil, result_master_fraction, 
result_master_ID,result_master_FD_fraction, result_master_C_in_FD_fraction, \
result_master_C_in_FD_WP, result_master_cem_fraction, run_parameters):
    global sr, reg_order, interval
    
    from openpyxl import Workbook
    import os
    import time;
    import sys
#    import pickle
    
    current_dir=os.getcwd()
#    file_list=pickle.load(open(current_dir+'/working_directory/file_list_conditioned.pkl','rb'))
    A=os.listdir(current_dir)
    if 'Results' not in A:
        if sys.platform[:3]=='win':
            os.makedirs(current_dir+'\\Results')
        else:
            os.makedirs(current_dir+'/Results')
    
    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    ws = wb.create_sheet(title="Run parameters")
#
#
#    ws.cell(row=1, column=1, value="sr")
#    ws.cell(row=1, column=2, value=run_summary["sr"])  
#    
#    ws.cell(row=2, column=1, value="reg_order")
#    ws.cell(row=2, column=2, value=run_summary["reg_order"]) 
#          
#    ws.cell(row=3, column=1, value="interval")
#    ws.cell(row=3, column=2, value=run_summary["interval"])  

    i=0
    keys=["sr","reg_order","interval","end_fit_WF","overal_fit_WF","err_end_slope_WF",\
    "err_maximum_transformation_WF",'Bs_model',"a0_gama",\
    "a0_alpha","CTE_alpha_a","CTE_alpha_b","CTE_alpha_c","c_wf_for_cte","c_wf_for_a0"]   
    for key in keys:
        i=i+1
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=run_parameters[key])  
   
    for i in xrange(len(file_list)):
        ws = wb.create_sheet(title=file_list[i][0])
        col=1
        ws.cell(row=1,column=col,value="time_conditioned(s)")    
        for row in xrange(2,len(dil_cond_master[i])+2):
            ws.cell(row=row, column=col, value=time_cond_master[i][row-2])
    
        col=2
        ws.cell(row=1,column=col,value="Temp_conditioned(C)")    
        for row in xrange(2,len(dil_cond_master[i])+2):
            ws.cell(row=row, column=col, value=temp_cond_master[i][row-2])
    
        col=3
        ws.cell(row=1,column=col,value="dil_conditioned(m)")    
        for row in xrange(2,len(dil_cond_master[i])+2):
            ws.cell(row=row, column=col, value=dil_cond_master[i][row-2])
    
        col=6
        ws.cell(row=1,column=col,value="time_result(s)")    
        for row in xrange(2,len(result_master_time[i])+2):
            ws.cell(row=row, column=col, value=result_master_time[i][row-2])
    
        col=7
        ws.cell(row=1,column=col,value="Temp_result(C)")    
        for row in xrange(2,len(result_master_temp[i])+2):
            ws.cell(row=row, column=col, value=result_master_temp[i][row-2])
    
        col=8
        ws.cell(row=1,column=col,value="dil_result(m)")    
        for row in xrange(2,len(result_master_dil[i])+2):
            ws.cell(row=row, column=col, value=result_master_dil[i][row-2])
            
        col=9
        ws.cell(row=1,column=col,value="fraction transformed")    
        for row in xrange(2,len(result_master_fraction[i])+2):
            ws.cell(row=row, column=col, value=result_master_fraction[i][row-2])

        col=10
        ws.cell(row=1,column=col,value="phase ID")    
        for row in xrange(2,len(result_master_ID[i])+2):
            ws.cell(row=row, column=col, value=result_master_ID[i][row-2])
    
        col=11
        ws.cell(row=1,column=col,value="fraction FD formed")    
        for row in xrange(2,len(result_master_FD_fraction[i])+2):
            ws.cell(row=row, column=col, value=result_master_FD_fraction[i][row-2])
            
        col=12
        ws.cell(row=1,column=col,value="C in FD(mole fraction)")    
        for row in xrange(2,len(result_master_C_in_FD_fraction[i])+2):
            ws.cell(row=row, column=col, value=result_master_C_in_FD_fraction[i][row-2])
            
        col=13
        ws.cell(row=1,column=col,value="C in FD(W%)")    
        for row in xrange(2,len(result_master_C_in_FD_WP[i])+2):
            ws.cell(row=row, column=col, value=result_master_C_in_FD_WP[i][row-2])
         
         
        col=14
        ws.cell(row=1,column=col,value="Fe in cementite(mole fraction)")    
        for row in xrange(2,len(result_master_cem_fraction[i])+2):
            ws.cell(row=row, column=col, value=result_master_cem_fraction[i][row-2])
            
#    localtime = time.asctime( time.localtime(time.time()) ).replace(':','-')
    wintime=time.strftime("%Y.%m.%d - %H.%M")

    Result_file_name=wintime+".xlsx"
    if sys.platform[:3]=='win':
        wb.save(current_dir+'\\Results\\'+Result_file_name)
    else:
        wb.save(current_dir+'/Results/'+Result_file_name)