# -*- coding: utf-8 -*-
"""
Created on Wed Dec 30 09:07:26 2015

Cross platform (Windows and Linux compatible)

@author: nasseh
"""
#This version is base on Results_xlsx_v4 in the old version control system.

from collections import OrderedDict
global sr, reg_order, interval
#this sript saves the results of calculation in the xlsx format
def saver(file_list,time_cond_master, temp_cond_master, dil_cond_master,
result_master_time, result_master_temp, result_master_dil, result_master_fraction, result_master_ID,result_master_FD_fraction, result_master_C_in_FD_fraction, result_master_C_in_FD_WP, result_master_cem_fraction,result_master_Vol_f_transformed, result_master_FD_vol,result_master_cem_vol,result_master_aus_vol, run_parameters):
    global sr, reg_order, interval
    
    from openpyxl import Workbook
    import os
    import time;
    import sys
#    import pickle
    
    current_dir=os.getcwd()
#    file_list=pickle.load(open(current_dir+'/working_directory/file_list_conditioned.pkl','rb'))
    A=os.listdir(current_dir)
    if 'Results' not in A:
        if sys.platform[:3]=='win':
            os.makedirs(current_dir+'\\Results')
        else:
            os.makedirs(current_dir+'/Results')
    
    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    ws = wb.create_sheet(title="Run parameters")

    i=0
    keys=["script_name","sr","reg_order","interval","optimized","L0_Correction",    
    "end_fit_WF","overal_fit_WF","err_end_slope_WF","err_maximum_transformation_WF",
    'Bs_model', "Ms_model","a0_gama","a0_alpha","CTE_alpha_a","CTE_alpha_b","CTE_alpha_c", "c_wf_for_cte","c_wf_for_a0","use avr CTE product for initial quess of CTE_alpha"]   
    for key in keys:
        i=i+1
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=run_parameters[key])  
        
    Columns=OrderedDict([("Time_conditioned(s)",time_cond_master),
              ("Temp_conditioned(C)", temp_cond_master),
              ("dil_conditioned(m)", dil_cond_master),
              ("time_result(s)", result_master_time),
              ("Temp_result(C)", result_master_temp),
              ("dil_result(m)", result_master_dil),
              ("Mole fraction transformed", result_master_fraction),
              ("Volume fraction transformed", result_master_Vol_f_transformed),
              ("phase ID", result_master_ID),
              ("Mole fraction FD formed", result_master_FD_fraction),
              ("Volume fraction FD formed",result_master_FD_vol),
              ("C in FD(mole fraction)", result_master_C_in_FD_fraction),
              ("C in FD(W%)", result_master_C_in_FD_WP),
              ("Fe in cementite(mole fraction)", result_master_cem_fraction)])
              
              
    #, "Temp_conditioned(C)","dil_conditioned(m)", "time_result(s)", "Temp_result(C)","dil_result(m)","fraction transformed","phase ID", "fraction FD formed",}
    for i in range(len(file_list)):
        ws = wb.create_sheet(title=file_list[i][0])
        
        for j in range(len(Columns.keys())):
            #print len(Collumns)
            ws.cell(row=1,column=j+1,value=list(Columns.keys())[j])    
            for row in range(2,len(Columns[list(Columns.keys())[j]][i])+2):
                ws.cell(row=row, column=j+1, value=Columns[list(Columns.keys())[j]][i][row-2])
        
    wintime=time.strftime("%Y.%m.%d - %H.%M")

    Result_file_name=wintime+".xlsx"
    if sys.platform[:3]=='win':
        wb.save(current_dir+'\\Results\\'+Result_file_name)
    else:
        wb.save(current_dir+'/Results/'+Result_file_name)

    #print "***888****8888****"
    #print (current_dir+'/Results/'+Result_file_name)
    file_location=str('./Results/'+Result_file_name)
    return file_location
    #call([ 'open file_location'])
#    call(['open', file_location])
#    os.sys('open 12.xlsx')